import os
import sys
import errno
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Alignment, Font, Border, Color, PatternFill

current_path = sys.path[0]

file_name = current_path + '\\data\\Store_data.xlsx'
TAX_RATE = 1.05
SALES_TAX = 0.00


class InvoiceController:
    def __init__(self, xl_file):
        if(os.path.isfile(xl_file)):
            self.xl_file = xl_file
        else:
            raise FileNotFoundError(
                errno.ENOENT, os.strerror(errno.ENOENT), xl_file)
        self.customers = {}
        self.invoices = {}
        self.line_items = {}
        self.products = {}
        self.wb = load_workbook(self.xl_file)

    def loadCustomer(self):
        ws = self.wb['customers']
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
            data = {"customer_id": int(row[0]), "first_name": row[1],
                    "last_name": row[2],  "address": row[3]}
            self.customers[row[0]] = data

    def loadProducts(self):
        ws = self.wb['products']
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
            data = {"product_id": int(row[0]), "description": row[1],
                    "price": float(row[2])}
            self.products[row[0]] = data

    def loadInvoices(self):
        ws = self.wb['invoices']
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
            data = {"invoice_id": int(row[0]), "customer_id": int(row[1]),
                    "invoice_date": row[2]}
            self.invoices[row[0]] = data

    def loadLineItems(self):
        ws = self.wb['line_items']
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
            data = {"line_item_id": row[0], "invoice_id": row[1],
                    "product_id": row[2], "qty": int(row[3])}
            self.line_items[row[0]] = data

    def ListInvoices(self):
        print('Available invoices: ')
        for i in self.invoices:
            print(str(self.invoices[i]["invoice_id"]) + ", ", end=" ")

    def getInvoice(self, invoice_id):
        return self.invoices[invoice_id]

    def getCustomer(self, customer_id):
        return self.customers[customer_id]

    def getLineItem(self, line_item_id):
        return self.line_items[line_item_id]

    def getProduct(self, product_id):
        return self.products[product_id]

    def buildLineItem(self, invoiceNumber):
        line_items = {}
        for row in self.line_items:
            if self.line_items[row]["invoice_id"] == invoiceNumber:
                data = self.getProduct(self.line_items[row]["product_id"])
                line_item = self.getLineItem(row)
                data.update(line_item)
                line_items[row] = data
                total = {"total": round(line_items[row]
                                        ['price'] * line_items[row]['qty'], 2)}
                line_items[row].update(total)
        return line_items

    def printInvoice(self, invoiceNumber):
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('B1:G1')
        ws['B1'].value = "Joe's Hardware"

        ws.merge_cells('B2:C2')
        ws['B2'].value = "123 Some Street"
        ws.merge_cells('D2:E2')
        ws['D2'].value = "P: 403-555-1234"
        ws.merge_cells('D3:E3')
        ws['D3'].value = "F: 403-555-5678"

        ws.merge_cells('F2:G2')
        ws['F2'].value = "joe.hardware@hardware.ca"

        ws.merge_cells('B3:C3')
        ws['B3'].value = "Calgary, Alberta, T1E 4M3"

        ws.merge_cells('F3:G3')
        ws['F3'].value = "http://hardware.ca"

        ws['B4'].value = "Bill To:"
        ws.merge_cells('B5:B6')
        ws['B5'].value = "Address:"

        ws.merge_cells('C5:C6')
        ws["C5"].alignment = Alignment(wrap_text=True)

        ws.merge_cells('D4:E6')

        ws['F4'].value = "Invoice#:"
        ws['F5'].value = "Invoice Date:"
        ws['B8'].value = "Item #"
        ws['C8'].value = "Description"
        ws['D8'].value = "Qty"
        ws['E8'].value = "Unit Price"
        ws['F8'].value = "Discount"
        ws['G8'].value = "Price"
        ws['F21'].value = "Invoice Subtotal"
        ws['F22'].value = "Tax Rate"
        ws['F23'].value = "Sales Tax"
        ws['F24'].value = "Other"
        ws['F25'].value = "Deposit Received"
        ws['F26'].value = "Total"

        invoice = self.getInvoice(invoiceNumber)
        customer = self.getCustomer(invoice['customer_id'])
        line_items = self.buildLineItem(invoiceNumber)
        print(line_items)
        ws.title = str(customer['first_name']+" " + customer['last_name'])

        ws['C4'].value = ws.title
        ws['C5'].value = customer['address']
        ws['G4'].value = invoice['invoice_id']
        ws['G5'].value = invoice['invoice_date']
        count = 9
        for line in line_items:
            ws['B'+str(count)].value = line_items[line]['product_id']
            ws['C'+str(count)].value = line_items[line]['description']
            ws['D'+str(count)].value = line_items[line]['qty']
            ws['E'+str(count)].value = line_items[line]['price']
            ws['G'+str(count)].value = line_items[line]['total']
            count += 1

        ws['G21'] = "=SUM(G9:G20)"
        ws['G21'].number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
        ws['G22'].value = TAX_RATE
        ws['G26'] = "= (G21*G22)+G24-G25"
        ws['G26'].number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'

        wb.save(current_path+"/invoice.xlsx")


def main():
    ic = InvoiceController(file_name)
    ic.loadCustomer()
    ic.loadProducts()
    ic.loadInvoices()
    ic.loadLineItems()
    while True:
        ic.ListInvoices()
        answer = input(
            "Which invoice number do you want to print? or \033[4mE\033[0mxit ")
        if(not answer.isdigit()):
            if(answer.lower() == "exit" or "e"):
                break
        else:
            answer = int(answer)
            ic.printInvoice(answer)


if __name__ == '__main__':
    main()
