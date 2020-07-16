import os
import sys
import asyncio
from invoices import InvoiceController
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

import pytest


current_path = sys.path[0]
file_name = current_path + '\\data\\Store_data.xlsx'
invoice = current_path+"\\invoice.xlsx"
ic = InvoiceController(file_name)


@pytest.mark.parametrize("function, dictionary, result", [
    (ic.loadCustomer(), ic.customers, 0),
    (ic.loadInvoices(), ic.invoices, 0),
    (ic.loadLineItems(), ic.line_items, 0),
    (ic.loadProducts(), ic.products, 0)
])
def test_load_database(function, dictionary, result):
    function            # setup
    count = len(dictionary)  # action
    assert count > result  # assert


@pytest.mark.parametrize("product_id, description, price", [
    (9, 'Bathroom & Kitchen Cleaner', 7.99),
    (10, 'Claw Hammer, 16-oz', 7.99),
    (11, 'Majic Eraser, 2pk', 2.99)
])
def test_get_product(product_id, description, price):
    ic.loadProducts()
    product = ic.getProduct(product_id)  # action
    assert product["product_id"] == product_id  # assert
    assert product["description"] == description  # assert
    assert product["price"] == price  # assert


@pytest.mark.parametrize("customer_id, first_name,last_name, address", [
    (1, 'Amalee', 'Sapsford', '37907 Upham Hill'),
    (16, 'Hernando', 'MacCafferty', '016 Rockefeller Place'),
    (18, 'Basilius', 'Dougher', '4 Fordem Lane'),
    (20, 'Consuelo', 'Attle', '5990 Thackeray Lane'),
])
def test_get_customer(customer_id, first_name, last_name, address):
    ic.loadCustomer()
    customer = ic.getCustomer(customer_id)  # action
    assert customer["customer_id"] == customer_id  # assert
    assert customer["first_name"] == first_name  # assert
    assert customer["last_name"] == last_name  # assert
    assert customer["address"] == address  # assert


# 'line_item_id': 82, 'invoice_id': 12, 'product_id': 9, 'qty': 15
@pytest.mark.parametrize("line_item_id, invoice_id,product_id, qty", [
    (82,  12, 9, 15),
    (23, 45, 20, 11),
    (174, 88, 18, 1),
    (174, 88, 18, 1),
])
def test_get_lineItem(line_item_id, invoice_id, product_id, qty):
    ic.loadLineItems()
    line = ic.getLineItem(line_item_id)  # action
    assert line["line_item_id"] == line_item_id  # assert
    assert line["invoice_id"] == invoice_id  # assert
    assert line["product_id"] == product_id  # assert
    assert line["qty"] == qty  # assert


def test_buildLineItem():
    ic.loadProducts()
    ic.loadLineItems()
    line_items = ic.buildLineItem(34)  # action
    assert len(line_items) > 0


@pytest.fixture
async def test_printInvoice():
    ic.loadCustomer()
    ic.loadInvoices()
    ic.loadLineItems()
    ic.loadProducts()
    ic.printInvoice(23)
    await asyncio.sleep(1000)
    wb = load_workbook(file_name)
    ws = wb.active
    result = ws['B2'].value
    assert result == 23
